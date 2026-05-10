# Importowanie wymaganych bibliotek
import pandas as pd # Do pracy z ramkami danych
import numpy as np # Do operacji numerycznych
from pathlib import Path # Do operacji na ścieżkach plików
import json # Do pracy z formatem JSON
import logging # Do logowania informacji
from datetime import datetime # Do pracy z datami i czasem

# Konfiguracja systemu logowania
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

# Próba importu biblioteki openpyxl do obsługi plików Excel
try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl nie zainstalowany - XLSX z filtrami nie beda generowane << dorzuc openpyxl do requirements.txt i zainstaluj, aby miec te pliki >>")

# Definicja klasy do eksportu danych
class DataExporter:
    # Inicjalizacja obiektu DataExporter
    def __init__(self):
        self.df = None # Główna ramka danych
        self.train_df = None # Ramka danych dla zbioru treningowego
        self.val_df = None # Ramka danych dla zbioru walidacyjnego
        self.test_df = None # Ramka danych dla zbioru testowego
        self.export_info = {} # Słownik do przechowywania informacji o eksporcie
    
    # Metoda do ładowania danych po inżynierii cech
    def load_data(self):
        # Ustalenie ścieżki do katalogu 'data'
        data_dir = Path(__file__).parent / "data"
        # Definiowanie ścieżki do pliku z danymi po inżynierii cech
        input_file = data_dir / "games_engineered.csv"
        # Wczytanie danych do DataFrame
        self.df = pd.read_csv(input_file, index_col=False)
        logger.info(f"Zaladowano dane: {self.df.shape}")
        return self.df
    
    # Metoda do selekcji finalnych cech do modelowania
    def select_final_features(self):
        logger.info("Wybieranie 15 kluczowych cech do modelowania...")
        # Lista kolumn, które mają zostać zachowane
        kept_columns = [
            'AppID', 'Name', 'Genres',
            'Release_year', 'Days_since_release',
            'Platform_count', 'Price', 'Is_free',
            'Total_reviews', 'Review_ratio', 'Is_highly_rated',
            'Log_owners', 'Has_achievements', 'Log_total_reviews', 'Genre_count'
        ]
        # Filtrowanie kolumn, aby zachować tylko te, które istnieją w DataFrame
        final_columns = [col for col in kept_columns if col in self.df.columns]
        # Aktualizacja DataFrame o wybrane kolumny
        self.df = self.df[final_columns].copy()
        logger.info(f"OK Wybrano {len(final_columns)} cech")
        logger.info(f"Ostateczne dane: {self.df.shape}")
        return self.df
    
    # Metoda do eksportu danych do formatu CSV
    def export_csv(self):
        logger.info("Eksportowanie do CSV...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        # Definiowanie nazwy pliku wyjściowego
        output_file = output_dir / "games_final.csv"
        # Zapisanie DataFrame do pliku CSV
        self.df.to_csv(output_file, index=False)
        logger.info(f"OK Eksportowano: {output_file.name}")
        self.export_info['csv'] = str(output_file) # Zapisanie ścieżki do informacji o eksporcie
        return self.df
    
    # Metoda do eksportu danych do formatu Parquet
    def export_parquet(self):
        logger.info("Eksportowanie do Parquet...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        # Definiowanie nazwy pliku wyjściowego
        output_file = output_dir / "games_final.parquet"
        # Zapisanie DataFrame do pliku Parquet z kompresją gzip
        self.df.to_parquet(output_file, index=False, compression='gzip')
        logger.info(f"OK Eksportowano: {output_file.name}")
        self.export_info['parquet'] = str(output_file) # Zapisanie ścieżki do informacji o eksporcie
        return self.df
    
    # Metoda do tworzenia podziału danych na zbiory treningowy, walidacyjny i testowy
    def create_train_val_test_split(self, train_ratio=0.7, val_ratio=0.15, test_ratio=0.15, random_state=42):
        logger.info(f"Tworzenie train/val/test splitu (train={train_ratio*100:.0f}%/val={val_ratio*100:.0f}%/test={test_ratio*100:.0f}%)...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)

        # Tasowanie ramki danych i resetowanie indeksu
        df_shuffled = self.df.sample(frac=1, random_state=random_state).reset_index(drop=True)
        total_len = len(df_shuffled) # Całkowita długość ramki danych
        train_end = int(total_len * train_ratio) # Indeks końca zbioru treningowego
        val_end = train_end + int(total_len * val_ratio) # Indeks końca zbioru walidacyjnego

        # Podział ramki danych na trzy zbiory
        self.train_df = df_shuffled.iloc[:train_end]
        self.val_df = df_shuffled.iloc[train_end:val_end]
        self.test_df = df_shuffled.iloc[val_end:]

        # Definiowanie ścieżek do plików wyjściowych
        train_file = output_dir / "games_train.csv"
        val_file = output_dir / "games_val.csv"
        test_file = output_dir / "games_test.csv"
        # Zapisywanie zbiorów do plików CSV
        self.train_df.to_csv(train_file, index=False)
        self.val_df.to_csv(val_file, index=False)
        self.test_df.to_csv(test_file, index=False)
        logger.info(f"OK Train set: {len(self.train_df)} wierszy ({len(self.train_df)/total_len*100:.1f}%)")
        logger.info(f"OK Validation set: {len(self.val_df)} wierszy ({len(self.val_df)/total_len*100:.1f}%)")
        logger.info(f"OK Test set: {len(self.test_df)} wierszy ({len(self.test_df)/total_len*100:.1f}%)")
        # Zapisanie informacji o podziale do informacji o eksporcie
        self.export_info['train_val_test_split'] = {
            'train_file': str(train_file),
            'val_file': str(val_file),
            'test_file': str(test_file),
            'train_size': len(self.train_df),
            'val_size': len(self.val_df),
            'test_size': len(self.test_df),
            'split_ratios': {'train': train_ratio, 'val': val_ratio, 'test': test_ratio}
        }
        return self.train_df, self.val_df, self.test_df
    
    # Metoda do tworzenia grup cech
    def create_feature_groups(self):
        logger.info("Tworzenie grup cech dla 15 kolumn...")
        # Definicja grup cech
        feature_groups = {
            'identifiers': ['AppID', 'Name'],
            'temporal': ['Release_year', 'Days_since_release'],
            'platform': ['Platform_count'],
            'reviews': ['Total_reviews', 'Review_ratio', 'Log_total_reviews'],
            'scores': ['Is_highly_rated'],
            'content': ['Log_owners', 'Has_achievements', 'Genre_count'],
            'price': ['Price', 'Is_free'],
            'metadata': ['Genres']
            #podział na gatunki i który jest najpopularniejszy
        }
        self.export_info['feature_groups'] = feature_groups # Zapisanie grup cech do informacji o eksporcie
        logger.info(f"OK Zdefiniowano {len(feature_groups)} grup cech")
        return feature_groups
    
    # Metoda do tworzenia dokumentacji kolumn
    def create_data_documentation(self):
        logger.info("Tworzenie dokumentacji...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        # Słownik do przechowywania danych dokumentacji kolumn
        columns_doc = {'column_name': [], 'data_type': [], 'description': [], 'missing_count': [], 'unique_values': []}
        # Opisy dla poszczególnych kolumn
        descriptions = {
            'AppID': 'Unikalny identyfikator gry w Steam',
            'Name': 'Nazwa gry',
            'Genres': 'Gatunki gry (separator: przecinek)',
            'Release_year': 'Rok wydania gry',
            'Days_since_release': 'Liczba dni od wydania gry',
            'Platform_count': 'Liczba platform (0-3: brak, tylko jedna, dwie, wszystkie)',
            'Price': 'Cena gry w USD',
            'Is_free': 'Czy gra jest darmowa (1=tak, 0=nie)',
            'Total_reviews': 'Calkowita liczba recenzji (pozytywne + negatywne)',
            'Review_ratio': 'Udzial recenzji pozytywnych do calkowitych',
            'Is_highly_rated': 'Czy gra ma wysoka ocene (1=Metacritic >= 75)',
            'Log_owners': 'Log transformacja liczby oszacowanych wlascicieli',
            'Has_achievements': 'Czy gra posiada osiagniecia (1=tak, 0=nie)',
            'Log_total_reviews': 'Log transformacja calkowitej liczby recenzji',
            'Genre_count': 'Liczba gatunkow, ktorych przydzie gra'
        }
        # Iteracja przez kolumny DataFrame w celu zebrania informacji
        for col in self.df.columns:
            columns_doc['column_name'].append(col)
            columns_doc['data_type'].append(str(self.df[col].dtype))
            columns_doc['description'].append(descriptions.get(col, ''))
            columns_doc['missing_count'].append(int(self.df[col].isna().sum()))
            columns_doc['unique_values'].append(int(self.df[col].nunique()))
        # Utworzenie DataFrame z dokumentacją kolumn
        doc_df = pd.DataFrame(columns_doc)
        # Definiowanie nazwy pliku dokumentacji
        doc_file = output_dir / "columns_documentation.csv"
        # Zapisanie dokumentacji do pliku CSV
        doc_df.to_csv(doc_file, index=False)
        logger.info(f"OK Dokumentacja kolumn: {doc_file.name}")
        return doc_df
    
    # Metoda do tworzenia manifestu datasetu
    def create_manifest(self):
        logger.info("Tworzenie manifestu...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        # Słownik zawierający metadane datasetu
        manifest = {
            'dataset_name': 'Steam Games Dataset - Preprocessed',
            'creation_date': datetime.now().isoformat(),
            'total_records': len(self.df),
            'total_features': len(self.df.columns),
            'data_shape': list(self.df.shape),
            'feature_groups': self.export_info.get('feature_groups', {}),
            'train_val_test_split': self.export_info.get('train_val_test_split', {}),
            'files': {
                'main_data': 'games_final.csv',
                'parquet_data': 'games_final.parquet',
                'train_data': 'games_train.csv',
                'val_data': 'games_val.csv',
                'test_data': 'games_test.csv',
                'columns_doc': 'columns_documentation.csv',
                'manifest': 'dataset_manifest.json'
            },
            'quality_metrics': {
                'null_values': int(self.df.isna().sum().sum()),
                'duplicate_rows': int(self.df.duplicated().sum()),
                'memory_usage_mb': float(self.df.memory_usage(deep=True).sum() / 1024 / 1024)
            },
            'column_summary': {
                'numeric': len(self.df.select_dtypes(include=[np.number]).columns),
                'categorical': len(self.df.select_dtypes(include=['object']).columns),
                'datetime': len(self.df.select_dtypes(include=['datetime64']).columns)
            }
        }
        # Definiowanie nazwy pliku manifestu
        manifest_file = output_dir / "dataset_manifest.json"
        # Zapisanie manifestu do pliku JSON
        with open(manifest_file, 'w', encoding='utf-8') as f:
            json.dump(manifest, f, indent=2, ensure_ascii=False)
        logger.info(f"OK Manifest: {manifest_file.name}")
        return manifest
    
    # Metoda do tworzenia pliku README dla datasetu
    def create_readme(self):
        logger.info("Tworzenie README...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        # Treść pliku README
        readme_content = f"""# Steam Games Dataset - Przetworzenie

## Opis
Przetworzony dataset gier ze Steam przygotowany do modelowania ML.

## Zawartosc
- games_final.csv - Ostateczne dane w formacie CSV
- games_final.parquet - Ostateczne dane w formacie Parquet (binarny)
- games_train.csv - Zbior treningowy (70%)
- games_val.csv - Zbior walidacyjny (15%)
- games_test.csv - Zbior testowy (15%)
- columns_documentation.csv - Dokumentacja wszystkich kolumn
- dataset_manifest.json - Manifest i metadata datasetu

## Statystyka
- Calkowite rekordy: {len(self.df):,}
- Calkowite cechy: {len(self.df.columns)}
- Wartosci brakujace: {self.df.isna().sum().sum():,}
- Duplikaty: {self.df.duplicated().sum():,}

## Ladownie danych w Pythonie
```python
import pandas as pd

df = pd.read_csv('games_final.csv')

df = pd.read_parquet('games_final.parquet')

val_df = pd.read_csv('games_val.csv')
train_df = pd.read_csv('games_train.csv')
test_df = pd.read_csv('games_test.csv')
```

---
Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        # Definiowanie nazwy pliku README
        readme_file = output_dir / "dataset_documentation.md"
        # Zapisanie treści do pliku README
        with open(readme_file, 'w', encoding='utf-8') as f:
            f.write(readme_content)
        logger.info(f"OK README: {readme_file.name}")
    
    # Metoda do zapisywania podsumowania eksportu
    def save_summary(self):
        # Ustalenie ścieżki do pliku podsumowania w katalogu 'reports'
        summary_file = Path(__file__).parent / "reports" / "03_export_summary.json"
        summary_file.parent.mkdir(parents=True, exist_ok=True) # Utworzenie katalogu, jeśli nie istnieje
        # Dodanie informacji o kształcie danych, kolumnach i czasie do słownika export_info
        self.export_info['dataset_shape'] = list(self.df.shape)
        self.export_info['columns'] = list(self.df.columns)
        self.export_info['timestamp'] = datetime.now().isoformat()
        # Zapisanie podsumowania do pliku JSON
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(self.export_info, f, indent=2, ensure_ascii=False)
        logger.info(f"OK Streszczenie exportu: {summary_file.name}")
    
    # Metoda do eksportu grup cech do osobnych plików CSV
    def export_feature_groups_csv(self):
        logger.info("Generowanie CSV dla grup cech...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Definicja grup cech (powtórzenie, ale potrzebne do tej funkcji)
        feature_groups = {
            'identifiers': ['AppID', 'Name'],
            'temporal': ['Release_year', 'Days_since_release'],
            'platform': ['Platform_count'],
            'reviews': ['Total_reviews', 'Review_ratio', 'Log_total_reviews'],
            'scores': ['Is_highly_rated'],
            'content': ['Log_owners', 'Has_achievements', 'Genre_count'],
            'price': ['Price', 'Is_free'],
            'metadata': ['Genres']
        }
        
        # Iteracja przez grupy cech i zapisywanie ich do osobnych plików CSV
        for group_name, columns in feature_groups.items():
            cols_present = [col for col in columns if col in self.df.columns] # Filtrowanie istniejących kolumn
            if cols_present:
                group_df = self.df[cols_present].copy() # Tworzenie DataFrame dla grupy
                group_file = output_dir / f"games_group_{group_name}.csv" # Definiowanie nazwy pliku
                group_df.to_csv(group_file, index=False) # Zapisywanie do CSV
                logger.info(f"  OK {group_name}: {group_file.name} ({len(cols_present)} kolumn)")
    
    # Metoda do eksportu danych do plików XLSX z filtrami
    def export_with_filters_xlsx(self):
        # Sprawdzenie, czy biblioteka openpyxl jest dostępna
        if not HAS_OPENPYXL:
            logger.warning("  SKIP openpyxl nie dostepny - pomijanie XLSX z filtrami")
            return
        
        logger.info("Generowanie XLSX z filtrami na naglowkach...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Eksport całej ramki danych z filtrami
        output_file = output_dir / "games_final_with_filters.xlsx"
        self._create_xlsx_with_filters(self.df, output_file, "Games Data")
        logger.info(f"  OK games_final_with_filters.xlsx")
        
        # Eksport zbioru treningowego z filtrami
        if self.train_df is not None:
            train_file = output_dir / "games_train_with_filters.xlsx"
            self._create_xlsx_with_filters(self.train_df, train_file, "Training Data")
            logger.info(f"  OK games_train_with_filters.xlsx ({len(self.train_df)} wierszy)")
        else:
            logger.warning("  SKIP Brak danych treningowych do eksportu XLSX.")

        # Eksport zbioru walidacyjnego z filtrami
        if self.val_df is not None:
            val_file = output_dir / "games_val_with_filters.xlsx"
            self._create_xlsx_with_filters(self.val_df, val_file, "Validation Data")
            logger.info(f"  OK games_val_with_filters.xlsx ({len(self.val_df)} wierszy)")
        else:
            logger.warning("  SKIP Brak danych walidacyjnych do eksportu XLSX.")
        
        # Eksport zbioru testowego z filtrami
        if self.test_df is not None:
            test_file = output_dir / "games_test_with_filters.xlsx"
            self._create_xlsx_with_filters(self.test_df, test_file, "Test Data")
            logger.info(f"  OK games_test_with_filters.xlsx ({len(self.test_df)} wierszy)")
        else:
            logger.warning("  SKIP Brak danych testowych do eksportu XLSX.")
    
    # Prywatna metoda pomocnicza do tworzenia plików XLSX z filtrami i zamrożonymi nagłówkami
    def _create_xlsx_with_filters(self, df, output_file, sheet_name="Data"):
        wb = Workbook() # Utworzenie nowego skoroszytu Excel
        ws = wb.active # Aktywacja pierwszego arkusza
        ws.title = sheet_name # Ustawienie tytułu arkusza
        
        # Wpisanie nagłówków kolumn i ich formatowanie
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            from openpyxl.styles import Font, PatternFill
            cell.font = Font(bold=True, color="FFFFFF") # Pogrubiona czcionka, biały kolor
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Niebieskie tło
        
        # Wpisanie danych z DataFrame do arkusza
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Formatowanie liczb zmiennoprzecinkowych
                if isinstance(value, float):
                    cell.number_format = '0.00'
                # Formatowanie liczb całkowitych (z wyjątkiem pierwszej kolumny, która może być ID)
                elif isinstance(value, int) and col_idx != 1:
                    cell.number_format = '0'
        
        # Ustawienie szerokości kolumn na podstawie długości nagłówka lub wartości domyślnej
        for col_idx, col_name in enumerate(df.columns, 1):
            width = max(len(str(col_name)), 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.freeze_panes = "A2" # Zamrożenie pierwszego wiersza (nagłówków)
        
        # Dodanie autofiltrów do nagłówków
        max_row = len(df) + 1
        max_col = len(df.columns)
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        
        wb.save(output_file) # Zapisanie pliku Excel
    
    # Metoda do eksportu danych do plików XLSX z grupami cech jako oddzielnymi arkuszami
    def export_grouped_xlsx(self):
        # Sprawdzenie, czy biblioteka openpyxl jest dostępna
        if not HAS_OPENPYXL:
            logger.warning("  SKIP openpyxl nie dostepny - pomijanie XLSX z grupami")
            return
        
        logger.info("Generowanie XLSX z grupami cech jako oddzielne arkusze...")
        # Ustalenie ścieżki do katalogu 'data/processed' i utworzenie go, jeśli nie istnieje
        output_dir = Path(__file__).parent / "data" / "processed"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Definicja grup cech
        feature_groups = {
            'Identifiers': ['AppID', 'Name'],
            'Temporal': ['Release_year', 'Days_since_release'],
            'Platform': ['Platform_count'],
            'Reviews': ['Total_reviews', 'Review_ratio', 'Log_total_reviews'],
            'Scores': ['Is_highly_rated'],
            'Content': ['Log_owners', 'Has_achievements', 'Genre_count'],
            'Price': ['Price', 'Is_free'],
            'Metadata': ['Genres']
        }
        
        output_file = output_dir / "games_final_grouped.xlsx" # Definiowanie nazwy pliku wyjściowego
        wb = Workbook() # Utworzenie nowego skoroszytu Excel
        wb.remove(wb.active) # Usunięcie domyślnego pustego arkusza
        
        # Iteracja przez grupy cech
        for group_name, columns in feature_groups.items():
            cols_present = [col for col in columns if col in self.df.columns] # Filtrowanie istniejących kolumn
            if cols_present:
                ws = wb.create_sheet(group_name) # Utworzenie nowego arkusza dla grupy
                group_df = self.df[cols_present].copy() # Tworzenie DataFrame dla grupy
                
                # Wpisanie nagłówków kolumn i ich formatowanie
                for col_idx, col_name in enumerate(cols_present, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    from openpyxl.styles import Font, PatternFill
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Wpisanie danych z DataFrame do arkusza
                for row_idx, row in enumerate(group_df.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if isinstance(value, float):
                            cell.number_format = '0.00'
                
                ws.freeze_panes = "A2" # Zamrożenie pierwszego wiersza (nagłówków)
                # Ustawienie szerokości kolumn
                for col_idx, col_name in enumerate(cols_present, 1):
                    width = max(len(str(col_name)), 12)
                    from openpyxl.utils import get_column_letter
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        wb.save(output_file) # Zapisanie pliku Excel
        logger.info(f"  OK games_final_grouped.xlsx ({len(feature_groups)} arkuszy)")
    
    # Metoda uruchamiająca cały proces eksportu danych
    def run(self):
        logger.info("\n" + "=" * 80)
        logger.info("EXPORT I PRZYGOTOWANIE DANYCH")
        logger.info("=" * 80 + "\n")
        self.load_data() # Załadowanie danych
        self.select_final_features() # Selekcja finalnych cech
        
        logger.info("\n>>> PLIKI GLOWNE (CSV/Parquet)")
        self.export_csv() # Eksport do CSV
        self.export_parquet() # Eksport do Parquet
        self.create_train_val_test_split() # Tworzenie podziału na zbiory treningowy, walidacyjny i testowy
        
        logger.info("\n>>> PLIKI Z GRUPAMI I FILTRAMI")
        self.export_feature_groups_csv() # Eksport grup cech do CSV
        self.export_with_filters_xlsx() # Eksport do XLSX z filtrami
        self.export_grouped_xlsx() # Eksport do XLSX z grupami cech jako arkuszami
        
        logger.info("\n>>> DOKUMENTACJA")
        self.create_feature_groups() # Tworzenie grup cech (dla manifestu)
        self.create_data_documentation() # Tworzenie dokumentacji kolumn
        self.create_manifest() # Tworzenie manifestu datasetu
        self.create_readme() # Tworzenie pliku README
        self.save_summary() # Zapisywanie podsumowania eksportu
        
        logger.info("\n" + "=" * 80)
        logger.info("OK EXPORT UKONCZNY")
        logger.info("=" * 80 + "\n")
        logger.info("Pliki wyjsciowe w: data/processed/\n")
        logger.info("PLIKI GLOWNE (niezmienione):")
        logger.info("  games_final.csv")
        logger.info("  games_final.parquet")
        logger.info("  games_train.csv")
        logger.info("  games_val.csv")
        logger.info("  games_test.csv")
        logger.info("\n GRUPY KOLUMN (nowe):")
        logger.info("  games_group_identifiers.csv")
        logger.info("  games_group_temporal.csv")
        logger.info("  games_group_platform.csv")
        logger.info("  games_group_reviews.csv")
        logger.info("  games_group_scores.csv")
        logger.info("  games_group_content.csv")
        logger.info("  games_group_price.csv")
        logger.info("  games_group_metadata.csv")
        logger.info("\n XLSX Z FILTRAMI (nowe):")
        if HAS_OPENPYXL:
            logger.info("  games_final_with_filters.xlsx (wszystkie dane)")
            logger.info("  games_train_with_filters.xlsx (70% treningowe)")
            logger.info("  games_val_with_filters.xlsx (15% walidacyjne)")
            logger.info("  games_test_with_filters.xlsx (15% testowe)")
            logger.info("  games_final_grouped.xlsx (8 arkuszy z grupami)")
        else:
            logger.info("openpyxl nie dostepny - XLSX nie wygenerowane << dodaj openpyxl do requirements.txt i zainstaluj, aby miec te pliki >>")
        logger.info("\n DOKUMENTACJA:")
        logger.info("  columns_documentation.csv")
        logger.info("  dataset_manifest.json")
        logger.info("  dataset_documentation.md")

# Główna funkcja programu
def main():
    exporter = DataExporter() # Utworzenie instancji klasy DataExporter
    exporter.run() # Uruchomienie procesu eksportu

# Sprawdzenie, czy skrypt jest uruchamiany bezpośrednio
if __name__ == "__main__":
    main()
